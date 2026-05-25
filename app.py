"""
SCHOOL INVENTORY LENDING SYSTEM - Phase 3
==========================================
New Features:
- Admin authentication (password-protected)
- Delete entries from UI (students, equipment, transactions)
- Section-wide cooldown penalty system (not money!)
- Configurable cooldown period (default: 1 day)
"""

from flask import Flask, render_template, request, jsonify, session, redirect, url_for
import sqlite3
from datetime import datetime, timedelta, time
from functools import wraps
import csv
import io
import os
import secrets

# ============================================================================
# TIMEZONE CONFIGURATION
# ============================================================================
# Change this to your timezone (default: Asia/Manila for Philippines UTC+8)
try:
    import pytz
    TIMEZONE = pytz.timezone('Asia/Manila')
    USE_TIMEZONE = True
except ImportError:
    # If pytz not installed, fall back to system time
    TIMEZONE = None
    USE_TIMEZONE = False
    print("⚠️  Warning: pytz not installed. Using system time. Install with: pip install pytz")

def get_local_time():
    """Get current time in configured timezone"""
    if USE_TIMEZONE:
        return datetime.now(TIMEZONE)
    else:
        return datetime.now()

def localize_datetime(dt_string):
    """Convert ISO datetime string to local timezone"""
    if dt_string is None:
        return None
    
    dt = datetime.fromisoformat(dt_string.replace('Z', '+00:00'))
    
    if USE_TIMEZONE:
        # If datetime is naive (no timezone), assume UTC
        if dt.tzinfo is None:
            dt = pytz.UTC.localize(dt)
        # Convert to local timezone
        return dt.astimezone(TIMEZONE)
    else:
        return dt

app = Flask(__name__)
# Generate a secret key for sessions (change this in production!)
app.secret_key = secrets.token_hex(16)

DATABASE = 'school_inventory.db'

# ============================================================================
# ADMIN AUTHENTICATION
# ============================================================================

# Default admin credentials (CHANGE THESE!)
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "admin123"  # Change this to something secure!

def login_required(f):
    """Decorator to require admin login for protected routes"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_logged_in' not in session:
            return jsonify({'success': False, 'message': 'Authentication required'}), 401
        return f(*args, **kwargs)
    return decorated_function


# ============================================================================
# COOLDOWN PENALTY SYSTEM
# ============================================================================

class CooldownPenalty:
    """
    Instead of money penalties, late returns result in section-wide cooldowns.
    
    Example:
    - Juan from Grade 10 Section A returns late
    - ALL students in Grade 10 Section A get 1-day cooldown
    - They cannot borrow until cooldown expires
    """
    
    DEFAULT_COOLDOWN_DAYS = 1  # Change this to adjust default cooldown
    
    @staticmethod
    def calculate_cooldown(due_date, return_date=None):
        """
        Calculate cooldown period based on how late the return was.
        
        Rules:
        - On time or early: No cooldown
        - 0-2 hours late: 1 day cooldown (default)
        - 2-6 hours late: 2 days cooldown
        - 6-24 hours late: 3 days cooldown
        - More than 1 day late: 7 days cooldown
        """
        if return_date is None:
            return_date = get_local_time()  # Use local time
        
        # If returned on time, no cooldown
        if return_date <= due_date:
            return 0
        
        # Calculate hours late
        time_diff = return_date - due_date
        hours_late = time_diff.total_seconds() / 3600
        
        # Determine cooldown based on severity
        if hours_late <= 2:
            return CooldownPenalty.DEFAULT_COOLDOWN_DAYS  # 1 day default
        elif hours_late <= 6:
            return 2  # 2 days
        elif hours_late <= 24:
            return 3  # 3 days
        else:
            return 7  # 1 week for very late returns
    
    @staticmethod
    def apply_section_cooldown(conn, grade_level, section, cooldown_days):
        """
        Apply cooldown to an entire section.
        
        All students in that grade/section will be unable to borrow
        until the cooldown expires.
        """
        cooldown_until = get_local_time() + timedelta(days=cooldown_days)  # Use local time
        
        # Remove timezone for database storage
        if hasattr(cooldown_until, 'tzinfo') and cooldown_until.tzinfo is not None:
            cooldown_until = cooldown_until.replace(tzinfo=None)
        
        # Get all students in the section
        students = conn.execute('''
            SELECT id FROM students 
            WHERE grade_level = ? AND section = ?
        ''', (grade_level, section)).fetchall()
        
        # Apply cooldown to each student
        for student in students:
            conn.execute('''
                INSERT OR REPLACE INTO student_cooldowns 
                (student_id, grade_level, section, cooldown_until, reason)
                VALUES (?, ?, ?, ?, ?)
            ''', (student['id'], grade_level, section, cooldown_until, f'Late return - {cooldown_days} day cooldown'))
        
        conn.commit()
        return len(students), cooldown_until
    
    @staticmethod
    def check_cooldown(conn, student_id):
        """
        Check if a student is currently under cooldown.
        Returns (is_on_cooldown, cooldown_info)
        """
        now = get_local_time()  # Use local time
        # Remove timezone for database comparison
        if hasattr(now, 'tzinfo') and now.tzinfo is not None:
            now = now.replace(tzinfo=None)
        
        cooldown = conn.execute('''
            SELECT * FROM student_cooldowns 
            WHERE student_id = ? AND cooldown_until > ?
            ORDER BY cooldown_until DESC
            LIMIT 1
        ''', (student_id, now)).fetchone()
        
        if cooldown:
            return True, dict(cooldown)
        return False, None


# ============================================================================
# RETURN POLICY (from Phase 2)
# ============================================================================

class ReturnPolicy:
    SAME_DAY_CUTOFF = time(18, 0)  # 6:00 PM
    
    POLICIES = {
        'same_day': {
            'name': 'Same Day Return',
            'description': 'Must be returned by 6:00 PM today',
            'days': 0,
            'cutoff_time': time(18, 0)
        },
        'overnight': {
            'name': 'Overnight',
            'description': 'Return by 6:00 PM next day',
            'days': 1,
            'cutoff_time': time(18, 0)
        },
        'weekly': {
            'name': 'Weekly',
            'description': 'Return within 7 days',
            'days': 7,
            'cutoff_time': time(23, 59)
        }
    }
    
    @staticmethod
    def calculate_due_date(policy_type='same_day'):
        now = get_local_time()  # Use local time instead of datetime.now()
        policy = ReturnPolicy.POLICIES.get(policy_type, ReturnPolicy.POLICIES['same_day'])
        
        due_date = now + timedelta(days=policy['days'])
        cutoff = policy['cutoff_time']
        
        due_date = due_date.replace(
            hour=cutoff.hour,
            minute=cutoff.minute,
            second=0,
            microsecond=0
        )
        
        # Remove timezone info for database storage
        if hasattr(due_date, 'tzinfo') and due_date.tzinfo is not None:
            due_date = due_date.replace(tzinfo=None)
        
        return due_date


# ============================================================================
# DATABASE FUNCTIONS
# ============================================================================

def get_db_connection():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """Initialize database with all tables including cooldowns"""
    conn = get_db_connection()
    
    # Students table
    conn.execute('''
        CREATE TABLE IF NOT EXISTS students (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            lrn TEXT UNIQUE NOT NULL,
            first_name TEXT NOT NULL,
            last_name TEXT NOT NULL,
            grade_level TEXT,
            section TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Student cooldowns table (NEW in Phase 3)
    conn.execute('''
        CREATE TABLE IF NOT EXISTS student_cooldowns (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL,
            grade_level TEXT NOT NULL,
            section TEXT NOT NULL,
            cooldown_until TIMESTAMP NOT NULL,
            reason TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (student_id) REFERENCES students (id)
        )
    ''')
    
    # Equipment types table
    conn.execute('''
        CREATE TABLE IF NOT EXISTS equipment_types (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            category TEXT,
            description TEXT,
            return_policy TEXT DEFAULT 'same_day',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Equipment items table
    conn.execute('''
        CREATE TABLE IF NOT EXISTS equipment_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            equipment_type_id INTEGER NOT NULL,
            asset_tag TEXT UNIQUE NOT NULL,
            status TEXT DEFAULT 'available',
            condition TEXT DEFAULT 'good',
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (equipment_type_id) REFERENCES equipment_types (id)
        )
    ''')
    
    # Transactions table (penalty_amount removed, cooldown_days added)
    conn.execute('''
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_id INTEGER NOT NULL,
            equipment_item_id INTEGER NOT NULL,
            borrow_date TIMESTAMP,
            due_date TIMESTAMP NOT NULL,
            return_date TIMESTAMP,
            status TEXT DEFAULT 'borrowed',
            cooldown_days INTEGER DEFAULT 0,
            notes TEXT,
            FOREIGN KEY (student_id) REFERENCES students (id),
            FOREIGN KEY (equipment_item_id) REFERENCES equipment_items (id)
        )
    ''')
    
    conn.commit()
    conn.close()
    print("✓ Database initialized successfully!")


def add_sample_data():
    """Add sample data"""
    conn = get_db_connection()
    
    students = conn.execute('SELECT COUNT(*) as count FROM students').fetchone()
    if students['count'] > 0:
        conn.close()
        return
    
    # Add sample students
    sample_students = [
        ('123456789012', 'Juan', 'Dela Cruz', 'Grade 10', 'Section A'),
        ('123456789013', 'Maria', 'Santos', 'Grade 11', 'Section B'),
        ('123456789014', 'Pedro', 'Garcia', 'Grade 12', 'Section A'),
        ('123456789015', 'Anna', 'Reyes', 'Grade 10', 'Section A'),
        ('123456789016', 'Carlos', 'Mendoza', 'Grade 10', 'Section A'),
    ]
    
    conn.executemany('''
        INSERT INTO students (lrn, first_name, last_name, grade_level, section)
        VALUES (?, ?, ?, ?, ?)
    ''', sample_students)
    
    # Add equipment types (no penalty rates needed anymore)
    equipment_types = [
        ('Scientific Calculator', 'Electronics', 'Casio scientific calculator', 'same_day'),
        ('Basketball', 'Sports', 'Official size basketball', 'same_day'),
        ('Laptop', 'Electronics', 'Dell laptop for computer class', 'same_day'),
        ('Projector', 'Electronics', 'Portable projector', 'same_day'),
        ('Jump Rope', 'Sports', 'Standard jump rope', 'same_day'),
    ]
    
    for eq_type in equipment_types:
        conn.execute('''
            INSERT INTO equipment_types (name, category, description, return_policy)
            VALUES (?, ?, ?, ?)
        ''', eq_type)
    
    # Get equipment types
    types = conn.execute('SELECT * FROM equipment_types').fetchall()
    
    # Add individual items
    for eq_type in types:
        type_name = eq_type['name']
        type_id = eq_type['id']
        
        if 'Calculator' in type_name:
            count = 20
        elif 'Basketball' in type_name:
            count = 10
        elif 'Laptop' in type_name:
            count = 5
        elif 'Projector' in type_name:
            count = 3
        elif 'Jump Rope' in type_name:
            count = 15
        else:
            count = 5
        
        for i in range(1, count + 1):
            asset_tag = f"{type_name} #{i}"
            conn.execute('''
                INSERT INTO equipment_items (equipment_type_id, asset_tag, status, condition)
                VALUES (?, ?, 'available', 'good')
            ''', (type_id, asset_tag))
    
    conn.commit()
    conn.close()
    print("✓ Sample data added successfully!")


# ============================================================================
# WEB ROUTES
# ============================================================================

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/kiosk')
def kiosk():
    return render_template('kiosk.html')

@app.route('/kiosk/usb')
def kiosk_usb():
    """Kiosk optimized for USB QR/Barcode scanners"""
    return render_template('kiosk_usb_scanner.html')

@app.route('/admin')
def admin():
    # Check if admin is logged in
    if 'admin_logged_in' not in session:
        return redirect(url_for('admin_login'))
    return render_template('admin.html')

@app.route('/admin/login')
def admin_login():
    return render_template('admin_login.html')


# ============================================================================
# ADMIN AUTHENTICATION API
# ============================================================================

@app.route('/api/admin/login', methods=['POST'])
def admin_login_api():
    """Admin login endpoint"""
    data = request.json
    username = data.get('username')
    password = data.get('password')
    
    if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        session['admin_logged_in'] = True
        return jsonify({'success': True, 'message': 'Login successful'})
    else:
        return jsonify({'success': False, 'message': 'Invalid credentials'})

@app.route('/api/admin/logout', methods=['POST'])
def admin_logout_api():
    """Admin logout endpoint"""
    session.pop('admin_logged_in', None)
    return jsonify({'success': True, 'message': 'Logged out'})


# ============================================================================
# API ENDPOINTS - Student Operations
# ============================================================================

@app.route('/api/student/<lrn>')
def get_student(lrn):
    """Get student info and check cooldown status"""
    conn = get_db_connection()
    student = conn.execute('SELECT * FROM students WHERE lrn = ?', (lrn,)).fetchone()
    
    if not student:
        conn.close()
        return jsonify({'success': False, 'message': 'Student not found'})
    
    # Check cooldown status
    is_on_cooldown, cooldown_info = CooldownPenalty.check_cooldown(conn, student['id'])
    
    # Get currently borrowed items
    borrowed_items = conn.execute('''
        SELECT 
            t.*,
            ei.asset_tag,
            et.name as equipment_name
        FROM transactions t
        JOIN equipment_items ei ON t.equipment_item_id = ei.id
        JOIN equipment_types et ON ei.equipment_type_id = et.id
        WHERE t.student_id = ? AND t.status = 'borrowed'
        ORDER BY t.due_date
    ''', (student['id'],)).fetchall()
    
    conn.close()
    
    return jsonify({
        'success': True,
        'student': dict(student),
        'borrowed_items': [dict(item) for item in borrowed_items],
        'is_on_cooldown': is_on_cooldown,
        'cooldown_info': cooldown_info
    })


@app.route('/api/equipment/types')
def get_equipment_types():
    """Get all equipment types with availability"""
    conn = get_db_connection()
    
    types = conn.execute('''
        SELECT 
            et.*,
            COUNT(CASE WHEN ei.status = 'available' THEN 1 END) as available_count,
            COUNT(ei.id) as total_count
        FROM equipment_types et
        LEFT JOIN equipment_items ei ON et.id = ei.equipment_type_id
        GROUP BY et.id
        HAVING available_count > 0
        ORDER BY et.category, et.name
    ''').fetchall()
    
    conn.close()
    
    result = []
    for t in types:
        type_dict = dict(t)
        policy = ReturnPolicy.POLICIES.get(t['return_policy'], ReturnPolicy.POLICIES['same_day'])
        type_dict['policy_description'] = policy['description']
        result.append(type_dict)
    
    return jsonify(result)


@app.route('/api/equipment/items/<int:type_id>')
def get_available_items(type_id):
    """Get available items for a specific type"""
    conn = get_db_connection()
    
    items = conn.execute('''
        SELECT ei.*, et.name as type_name, et.category, et.return_policy
        FROM equipment_items ei
        JOIN equipment_types et ON ei.equipment_type_id = et.id
        WHERE ei.equipment_type_id = ? AND ei.status = 'available'
        ORDER BY ei.asset_tag
    ''', (type_id,)).fetchall()
    
    conn.close()
    return jsonify([dict(item) for item in items])


@app.route('/api/borrow', methods=['POST'])
def borrow_items():
    """
    Process borrowing transaction with cooldown check.
    NEW: Students can only borrow ONE item at a time.
    """
    data = request.json
    student_id = data['student_id']
    item_ids = data['item_ids']
    policy_type = data.get('policy_type', 'same_day')
    # 1. Get the local time at the start of the borrow process
    now = get_local_time()

    # 2. Strip timezone info for consistent database storage
    if hasattr(now, 'tzinfo') and now.tzinfo is not None:
        now_naive = now.replace(tzinfo=None)
    else:
        now_naive = now
    conn = get_db_connection()
    
    try:
        # Check if student already has borrowed items
        existing_borrows = conn.execute('''
            SELECT COUNT(*) as count FROM transactions 
            WHERE student_id = ? AND status = 'borrowed'
        ''', (student_id,)).fetchone()
        
        if existing_borrows['count'] > 0:
            conn.close()
            return jsonify({
                'success': False,
                'message': '⚠️ You already have an item borrowed. Please return it before borrowing another.',
                'error_type': 'already_borrowed'
            })
        
        # Limit to ONE item at a time
        if len(item_ids) > 1:
            conn.close()
            return jsonify({
                'success': False,
                'message': '⚠️ You can only borrow ONE item at a time.',
                'error_type': 'too_many_items'
            })
        
        # Check if student is on cooldown
        is_on_cooldown, cooldown_info = CooldownPenalty.check_cooldown(conn, student_id)
        
        if is_on_cooldown:
            conn.close()
            cooldown_until = datetime.fromisoformat(cooldown_info['cooldown_until'])
            return jsonify({
                'success': False,
                'message': f'Your section is on cooldown until {cooldown_until.strftime("%Y-%m-%d %I:%M %p")}',
                'cooldown_info': cooldown_info
            })
        
        # Calculate due date
        due_date = ReturnPolicy.calculate_due_date(policy_type)
        
        # Process borrowing
        for item_id in item_ids:
            item = conn.execute(
                'SELECT * FROM equipment_items WHERE id = ? AND status = "available"',
                (item_id,)
            ).fetchone()
            
            if not item:
                conn.rollback()
                conn.close()
                return jsonify({
                    'success': False,
                    'message': f'Item ID {item_id} is not available'
                })
            
            # Create transaction
            conn.execute('''
                INSERT INTO transactions (student_id, equipment_item_id, borrow_date, due_date, status)
                VALUES (?, ?, ?, ?, 'borrowed')
        ''', (student_id, item_id, now_naive, due_date))
            
            # Update item status
            conn.execute('''
                UPDATE equipment_items SET status = 'borrowed' WHERE id = ?
            ''', (item_id,))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{len(item_ids)} item(s) borrowed successfully',
            'due_date': due_date.strftime('%Y-%m-%d %I:%M %p')
        })
        
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@app.route('/api/return', methods=['POST'])
def return_items():
    """Process equipment return with cooldown penalty"""
    data = request.json
    transaction_ids = data['transaction_ids']
    
    conn = get_db_connection()
    # Use timezone-naive datetime (no timezone info - works with SQLite)
    return_date = datetime.now()
    
    try:
        returned_items = []
        section_cooldowns = {}  # Track cooldowns per section
        
        for trans_id in transaction_ids:
            # Get transaction details with student info
            trans = conn.execute('''
                SELECT t.*, ei.asset_tag, s.grade_level, s.section
                FROM transactions t
                JOIN equipment_items ei ON t.equipment_item_id = ei.id
                JOIN students s ON t.student_id = s.id
                WHERE t.id = ?
            ''', (trans_id,)).fetchone()
            
            if not trans:
                continue
            
            # Parse due_date from database (timezone-naive string)
            due_date = datetime.fromisoformat(trans['due_date'])
            
            # Both dates are now timezone-naive, comparison works!
            cooldown_days = CooldownPenalty.calculate_cooldown(due_date, return_date)
            
            # Update transaction
            conn.execute('''
                UPDATE transactions 
                SET return_date = ?, status = 'returned', cooldown_days = ?
                WHERE id = ?
            ''', (return_date, cooldown_days, trans_id))
            
            # Update item status
            conn.execute('''
                UPDATE equipment_items SET status = 'available' 
                WHERE id = ?
            ''', (trans['equipment_item_id'],))
            
            # Track section cooldown
            if cooldown_days > 0:
                section_key = f"{trans['grade_level']}|{trans['section']}"
                if section_key not in section_cooldowns:
                    section_cooldowns[section_key] = cooldown_days
                else:
                    # Use the longer cooldown if multiple items
                    section_cooldowns[section_key] = max(section_cooldowns[section_key], cooldown_days)
            
            returned_items.append({
                'asset_tag': trans['asset_tag'],
                'cooldown_days': cooldown_days,
                'hours_late': (return_date - due_date).total_seconds() / 3600 if return_date > due_date else 0
            })
        
        # Apply section cooldowns
        cooldown_messages = []
        for section_key, cooldown_days in section_cooldowns.items():
            grade_level, section = section_key.split('|')
            affected_count, cooldown_until = CooldownPenalty.apply_section_cooldown(
                conn, grade_level, section, cooldown_days
            )
            cooldown_messages.append({
                'grade_level': grade_level,
                'section': section,
                'cooldown_days': cooldown_days,
                'cooldown_until': cooldown_until.strftime('%Y-%m-%d %I:%M %p'),
                'affected_students': affected_count
            })
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{len(returned_items)} item(s) returned',
            'returned_items': returned_items,
            'cooldowns_applied': cooldown_messages
        })
        
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})
        
        # Apply section cooldowns
        cooldown_messages = []
        for section_key, cooldown_days in section_cooldowns.items():
            grade_level, section = section_key.split('|')
            affected_count, cooldown_until = CooldownPenalty.apply_section_cooldown(
                conn, grade_level, section, cooldown_days
            )
            cooldown_messages.append({
                'grade_level': grade_level,
                'section': section,
                'cooldown_days': cooldown_days,
                'cooldown_until': cooldown_until.strftime('%Y-%m-%d %I:%M %p'),
                'affected_students': affected_count
            })
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{len(returned_items)} item(s) returned',
            'returned_items': returned_items,
            'cooldowns_applied': cooldown_messages
        })
        
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


# ============================================================================
# API ENDPOINTS - Admin Operations (Protected)
# ============================================================================

@app.route('/api/admin/stats')
@login_required
def get_admin_stats():
    """Get statistics for dashboard"""
    conn = get_db_connection()
    
    now = get_local_time()
    if hasattr(now, 'tzinfo') and now.tzinfo is not None:
        now = now.replace(tzinfo=None)
    
    stats = {
        'total_items': conn.execute('SELECT COUNT(*) as count FROM equipment_items').fetchone()['count'],
        'available_items': conn.execute('SELECT COUNT(*) as count FROM equipment_items WHERE status = "available"').fetchone()['count'],
        'borrowed_items': conn.execute('SELECT COUNT(*) as count FROM equipment_items WHERE status = "borrowed"').fetchone()['count'],
        'total_students': conn.execute('SELECT COUNT(*) as count FROM students').fetchone()['count'],
        'equipment_types': conn.execute('SELECT COUNT(*) as count FROM equipment_types').fetchone()['count'],
        'active_transactions': conn.execute('SELECT COUNT(*) as count FROM transactions WHERE status = "borrowed"').fetchone()['count'],
        'overdue_items': conn.execute('SELECT COUNT(*) as count FROM transactions WHERE status = "borrowed" AND due_date < ?', (now,)).fetchone()['count'],
        'active_cooldowns': conn.execute('SELECT COUNT(*) as count FROM student_cooldowns WHERE cooldown_until > ?', (now,)).fetchone()['count'],
        'total_transactions': conn.execute('SELECT COUNT(*) as count FROM transactions').fetchone()['count']
    }
    
    conn.close()
    return jsonify(stats)


@app.route('/api/admin/reports/transactions/export')
@login_required
def export_transactions():
    """Export transaction history to Excel or PDF"""
    export_format = request.args.get('format', 'excel')  # 'excel' or 'pdf'
    
    conn = get_db_connection()
    
    transactions = conn.execute('''
        SELECT 
            t.id,
            t.borrow_date,
            t.due_date,
            t.return_date,
            t.status,
            t.cooldown_days,
            s.lrn,
            s.first_name || ' ' || s.last_name as student_name,
            s.grade_level,
            s.section,
            ei.asset_tag,
            et.name as equipment_name,
            et.category
        FROM transactions t
        JOIN students s ON t.student_id = s.id
        JOIN equipment_items ei ON t.equipment_item_id = ei.id
        JOIN equipment_types et ON ei.equipment_type_id = et.id
        ORDER BY t.borrow_date DESC
    ''').fetchall()
    
    conn.close()
    
    if export_format == 'excel':
        return generate_excel_report(transactions)
    else:
        return generate_pdf_report(transactions)


def generate_excel_report(transactions):
    """Generate Excel file from transaction data"""
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return jsonify({
            'success': False,
            'message': 'openpyxl library not installed. Run: pip install openpyxl'
        }), 500
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Transaction History"
    
    # Headers
    headers = [
        'Transaction ID', 'Borrowed', 'Due Date', 'Return Date',
        'Status', 'Cooldown Days', 'Student Name', 'LRN', 'Grade', 'Section',
        'Equipment', 'Asset Tag', 'Category'
    ]
    
    # Style headers
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row_idx, trans in enumerate(transactions, 2):
        borrow_dt = datetime.fromisoformat(trans['borrow_date'])
        borrowed = borrow_dt.strftime('%Y-%m-%d %I:%M:%S %p')
        
        due_dt = datetime.fromisoformat(trans['due_date'])
        due_date = due_dt.strftime('%Y-%m-%d %I:%M:%S %p')
        
        return_date = datetime.fromisoformat(trans['return_date']).strftime('%Y-%m-%d %I:%M:%S %p') if trans['return_date'] else 'Not Returned'
        
        row_data = [
            trans['id'],
            borrowed,
            due_date,
            return_date,
            trans['status'],
            trans['cooldown_days'] if trans['cooldown_days'] > 0 else 'None',
            trans['student_name'],
            trans['lrn'],
            trans['grade_level'],
            trans['section'],
            trans['equipment_name'],
            trans['asset_tag'],
            trans['category']
        ]
        
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    from flask import send_file
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'transaction_history_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


def generate_pdf_report(transactions):
    """Generate PDF file from transaction data"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
    except ImportError:
        return jsonify({
            'success': False,
            'message': 'reportlab library not installed. Run: pip install reportlab'
        }), 500
    
    import io
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph("<b>Transaction History Report</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    # Metadata
    metadata = Paragraph(
        f"Generated: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}<br/>"
        f"Total Transactions: {len(transactions)}",
        styles['Normal']
    )
    elements.append(metadata)
    elements.append(Spacer(1, 0.3*inch))
    
    # Table data
    data = [[
        'ID', 'Borrowed', 'Returned', 'Student', 'Grade/Sec',
        'Equipment', 'Asset Tag', 'Cooldown'
    ]]
    
    for trans in transactions[:100]:  # Limit to 100 for PDF size
        borrow_dt = datetime.fromisoformat(trans['borrow_date'])
        borrowed = borrow_dt.strftime('%Y-%m-%d\n%I:%M %p')
        returned = datetime.fromisoformat(trans['return_date']).strftime('%Y-%m-%d %I:%M:%S %p') if trans['return_date'] else 'Not Returned'
        
        data.append([
            str(trans['id']),
            borrowed,
            returned,
            trans['student_name'][:20],
            f"{trans['grade_level']}\n{trans['section']}",
            trans['equipment_name'][:15],
            trans['asset_tag'][:15],
            f"{trans['cooldown_days']}d" if trans['cooldown_days'] > 0 else '-'
        ])
    
    # Create table
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(table)
    
    if len(transactions) > 100:
        elements.append(Spacer(1, 0.2*inch))
        note = Paragraph(
            f"<i>Note: Showing first 100 of {len(transactions)} transactions. "
            "Export to Excel for complete data.</i>",
            styles['Normal']
        )
        elements.append(note)
    
    doc.build(elements)
    buffer.seek(0)
    
    from flask import send_file
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'transaction_history_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    )


@app.route('/api/admin/transactions')
@login_required
def get_transactions():
    """Get all transactions with details"""
    conn = get_db_connection()
    
    transactions = conn.execute('''
        SELECT 
            t.*,
            s.lrn,
            s.first_name || ' ' || s.last_name as student_name,
            s.grade_level,
            s.section,
            ei.asset_tag,
            et.name as equipment_name
        FROM transactions t
        JOIN students s ON t.student_id = s.id
        JOIN equipment_items ei ON t.equipment_item_id = ei.id
        JOIN equipment_types et ON ei.equipment_type_id = et.id
        ORDER BY t.borrow_date DESC
        LIMIT 500
    ''').fetchall()
    
    conn.close()
    return jsonify([dict(t) for t in transactions])


@app.route('/api/admin/equipment/all')
@login_required
def get_all_equipment():
    """Get all equipment items"""
    conn = get_db_connection()
    
    items = conn.execute('''
        SELECT 
            ei.*,
            et.name as type_name,
            et.category,
            et.return_policy
        FROM equipment_items ei
        JOIN equipment_types et ON ei.equipment_type_id = et.id
        ORDER BY et.category, et.name, ei.asset_tag
    ''').fetchall()
    
    conn.close()
    return jsonify([dict(item) for item in items])


@app.route('/api/admin/students', methods=['POST'])
@login_required
def add_student_manual():
    """Manually add a single student through admin UI"""
    data = request.json
    
    conn = get_db_connection()
    
    try:
        conn.execute('''
            INSERT INTO students (lrn, first_name, last_name, grade_level, section)
            VALUES (?, ?, ?, ?, ?)
        ''', (
            data['lrn'],
            data['first_name'],
            data['last_name'],
            data.get('grade_level', ''),
            data.get('section', '')
        ))
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Student added successfully'
        })
        
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({
            'success': False,
            'message': 'Student with this LRN already exists'
        })
    except Exception as e:
        conn.close()
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        })


@app.route('/api/admin/equipment', methods=['POST'])
@login_required
def add_equipment_manual():
    """Manually add a single equipment item through admin UI"""
    data = request.json
    
    conn = get_db_connection()
    
    try:
        # Get equipment type ID
        eq_type = conn.execute(
            'SELECT id FROM equipment_types WHERE name = ?',
            (data['equipment_type'],)
        ).fetchone()
        
        if not eq_type:
            conn.close()
            return jsonify({
                'success': False,
                'message': f'Equipment type "{data["equipment_type"]}" not found'
            })
        
        conn.execute('''
            INSERT INTO equipment_items (equipment_type_id, asset_tag, status, condition, notes)
            VALUES (?, ?, 'available', ?, ?)
        ''', (
            eq_type['id'],
            data['asset_tag'],
            data.get('condition', 'good'),
            data.get('notes', '')
        ))
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Equipment added successfully'
        })
        
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({
            'success': False,
            'message': 'Equipment with this asset tag already exists'
        })
    except Exception as e:
        conn.close()
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        })


@app.route('/api/admin/delete-all/students', methods=['POST'])
@login_required
def delete_all_students():
    """Delete all students (for new school year)"""
    conn = get_db_connection()
    
    try:
        # Check for active borrows
        active = conn.execute(
            'SELECT COUNT(*) as count FROM transactions WHERE status = "borrowed"'
        ).fetchone()
        
        if active['count'] > 0:
            conn.close()
            return jsonify({
                'success': False,
                'message': f'Cannot delete: {active["count"]} items are currently borrowed. Return all items first.'
            })
        
        # Delete all students and their cooldowns
        conn.execute('DELETE FROM student_cooldowns')
        conn.execute('DELETE FROM transactions')
        count = conn.execute('DELETE FROM students').rowcount
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Deleted {count} students and all transaction history'
        })
        
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        })


@app.route('/api/admin/delete-all/equipment', methods=['POST'])
@login_required
def delete_all_equipment():
    """Delete all equipment items (for new school year)"""
    conn = get_db_connection()
    
    try:
        # Check for borrowed items
        borrowed = conn.execute(
            'SELECT COUNT(*) as count FROM equipment_items WHERE status = "borrowed"'
        ).fetchone()
        
        if borrowed['count'] > 0:
            conn.close()
            return jsonify({
                'success': False,
                'message': f'Cannot delete: {borrowed["count"]} items are currently borrowed. Return all items first.'
            })
        
        # Delete all equipment items
        count = conn.execute('DELETE FROM equipment_items').rowcount
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Deleted {count} equipment items'
        })
        
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        })


@app.route('/api/admin/delete-all/transactions', methods=['POST'])
@login_required
def delete_all_transactions():
    """Delete all transaction history (for new school year)"""
    conn = get_db_connection()
    
    try:
        count = conn.execute('DELETE FROM transactions WHERE status = "returned"').rowcount
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Deleted {count} returned transactions'
        })
        
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        })


@app.route('/api/admin/equipment-types')
@login_required
def get_equipment_types_list():
    """Get list of equipment types for dropdown"""
    conn = get_db_connection()
    types = conn.execute('SELECT name FROM equipment_types ORDER BY name').fetchall()
    conn.close()
    
    return jsonify([t['name'] for t in types])


@app.route('/api/admin/students')
@login_required
def get_all_students():
    """Get all students"""
    conn = get_db_connection()
    students = conn.execute('SELECT * FROM students ORDER BY last_name, first_name').fetchall()
    conn.close()
    return jsonify([dict(s) for s in students])


@app.route('/api/admin/cooldowns')
@login_required
def get_active_cooldowns():
    """Get all active cooldowns"""
    conn = get_db_connection()
    
    cooldowns = conn.execute('''
        SELECT 
            sc.*,
            s.lrn,
            s.first_name || ' ' || s.last_name as student_name
        FROM student_cooldowns sc
        JOIN students s ON sc.student_id = s.id
        WHERE sc.cooldown_until > ?
        ORDER BY sc.cooldown_until DESC
    ''', (datetime.now(),)).fetchall()
    
    conn.close()
    return jsonify([dict(c) for c in cooldowns])


# ============================================================================
# DELETE OPERATIONS (NEW in Phase 3)
# ============================================================================

@app.route('/api/admin/delete/student/<int:student_id>', methods=['DELETE'])
@login_required
def delete_student(student_id):
    """Delete a student (only if no active transactions)"""
    conn = get_db_connection()
    
    # Check for active transactions
    active_trans = conn.execute('''
        SELECT COUNT(*) as count FROM transactions 
        WHERE student_id = ? AND status = 'borrowed'
    ''', (student_id,)).fetchone()
    
    if active_trans['count'] > 0:
        conn.close()
        return jsonify({
            'success': False,
            'message': 'Cannot delete student with active borrowed items'
        })
    
    try:
        conn.execute('DELETE FROM students WHERE id = ?', (student_id,))
        conn.execute('DELETE FROM student_cooldowns WHERE student_id = ?', (student_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Student deleted'})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@app.route('/api/admin/delete/equipment/<int:item_id>', methods=['DELETE'])
@login_required
def delete_equipment(item_id):
    """Delete equipment item (only if available)"""
    conn = get_db_connection()
    
    item = conn.execute('SELECT status FROM equipment_items WHERE id = ?', (item_id,)).fetchone()
    
    if not item:
        conn.close()
        return jsonify({'success': False, 'message': 'Item not found'})
    
    if item['status'] != 'available':
        conn.close()
        return jsonify({
            'success': False,
            'message': 'Cannot delete borrowed equipment'
        })
    
    try:
        conn.execute('DELETE FROM equipment_items WHERE id = ?', (item_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Equipment deleted'})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@app.route('/api/admin/delete/transaction/<int:trans_id>', methods=['DELETE'])
@login_required
def delete_transaction(trans_id):
    """Delete transaction (only if returned)"""
    conn = get_db_connection()
    
    trans = conn.execute('SELECT status FROM transactions WHERE id = ?', (trans_id,)).fetchone()
    
    if not trans:
        conn.close()
        return jsonify({'success': False, 'message': 'Transaction not found'})
    
    if trans['status'] == 'borrowed':
        conn.close()
        return jsonify({
            'success': False,
            'message': 'Cannot delete active transaction. Return item first.'
        })
    
    try:
        conn.execute('DELETE FROM transactions WHERE id = ?', (trans_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Transaction deleted'})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@app.route('/api/admin/delete/cooldown/<int:cooldown_id>', methods=['DELETE'])
@login_required
def delete_cooldown(cooldown_id):
    """Manually remove a cooldown"""
    conn = get_db_connection()
    
    try:
        conn.execute('DELETE FROM student_cooldowns WHERE id = ?', (cooldown_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Cooldown removed'})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


# ============================================================================
# CSV IMPORT (from Phase 2)
# ============================================================================

@app.route('/api/admin/import/students', methods=['POST'])
@login_required
def import_students_csv():
    """Import students from CSV file"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file provided'})
    
    file = request.files['file']
    
    try:
        stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
        csv_reader = csv.DictReader(stream)
        
        conn = get_db_connection()
        added = 0
        skipped = 0
        
        for row in csv_reader:
            try:
                conn.execute('''
                    INSERT INTO students (lrn, first_name, last_name, grade_level, section)
                    VALUES (?, ?, ?, ?, ?)
                ''', (
                    row['lrn'],
                    row['first_name'],
                    row['last_name'],
                    row.get('grade_level', ''),
                    row.get('section', '')
                ))
                added += 1
            except sqlite3.IntegrityError:
                skipped += 1
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Imported {added} students, skipped {skipped} duplicates',
            'added': added,
            'skipped': skipped
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@app.route('/api/admin/import/equipment', methods=['POST'])
@login_required
def import_equipment_csv():
    """Import equipment from CSV file"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file provided'})
    
    file = request.files['file']
    
    try:
        stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
        csv_reader = csv.DictReader(stream)
        
        conn = get_db_connection()
        added = 0
        skipped = 0
        
        for row in csv_reader:
            try:
                eq_type = conn.execute(
                    'SELECT id FROM equipment_types WHERE name = ?',
                    (row['equipment_type'],)
                ).fetchone()
                
                if not eq_type:
                    continue
                
                conn.execute('''
                    INSERT INTO equipment_items (equipment_type_id, asset_tag, status, condition, notes)
                    VALUES (?, ?, 'available', ?, ?)
                ''', (
                    eq_type['id'],
                    row['asset_tag'],
                    row.get('condition', 'good'),
                    row.get('notes', '')
                ))
                added += 1
            except sqlite3.IntegrityError:
                skipped += 1
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'Imported {added} items, skipped {skipped} duplicates',
            'added': added,
            'skipped': skipped
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


# ============================================================================
# APPLICATION STARTUP
# ============================================================================

if __name__ == '__main__':
    init_db()
    add_sample_data()
    
    print("\n" + "="*70)
    print("🎓 SCHOOL INVENTORY LENDING SYSTEM - Phase 3")
    print("="*70)
    print("\n✨ New Features:")
    print("  • Admin authentication (password-protected)")
    print("  • Delete entries from UI (students, equipment, transactions)")
    print("  • Section-wide cooldown penalty system (not money!)")
    print(f"  • Default cooldown: {CooldownPenalty.DEFAULT_COOLDOWN_DAYS} day(s)")
    print("\n🔐 Admin Credentials:")
    print(f"  Username: {ADMIN_USERNAME}")
    print(f"  Password: {ADMIN_PASSWORD}")
    print("  ⚠️  CHANGE THESE IN app.py (lines 28-29)")
    print("\n⏰ Cooldown Rules:")
    print("  • Late return: Entire section gets cooldown")
    print("  • 0-2 hours late: 1 day cooldown")
    print("  • 2-6 hours late: 2 days cooldown")
    print("  • 6-24 hours late: 3 days cooldown")
    print("  • >1 day late: 7 days cooldown")
    print("\nServer starting...")
    print("📱 Student Kiosk: http://localhost:5000/kiosk")
    print("👨‍💼 Admin Dashboard: http://localhost:5000/admin")
    print("\nPress CTRL+C to stop the server")
    print("="*70 + "\n")
    
    app.run(debug=True, host='127.0.0.1', port=5000)
